import base64
import io
from openpyxl import load_workbook
import xlsxwriter
from odoo import api, fields, models, _
from odoo.exceptions import ValidationError, UserError

from datetime import datetime, timedelta


class Operation_Deb(models.Model):
    _name = 'credit.operation.deb'
    _inherit = ["mail.thread", 'mail.activity.mixin']
    _description = "Opération Déblocage"

    name = fields.Char(string='Opération Déblocage', required=True, copy=False, readonly=True,
                       default=lambda self: _('New'))
    type_ligne = fields.Selection([('1', 'Crédit d\'exploitation'),
                                   ('2', 'Crédit d\'investissement'),
                                   ('3', 'Leasing'),
                                   ('4', 'Leasing adossé')], )
    state = fields.Selection([('draft', 'Prévisionnel'),
                              ('confirmed', 'Confirmé'),
                              ('extended', 'Prolongé')], default='draft')
    montant_debloque = fields.Float(string="Montant débloqué", store=True, required=True)
    montant_add = fields.Float(string="Interet", store=True)
    montant_total = fields.Float(string="Montant Total",)
    montant_total_comp = fields.Float(string="Montant Total", store=True, compute='_compute_total')
    montant_rembourser = fields.Float(string="Montant a rembourser")
    montant_remb_comp = fields.Float(string="Montant a rembourser", store=True, compute='_compute_reste')
    deblocage_date = fields.Date("Date de déblocage", tracking=True,)
    date_prevue_deblocage = fields.Date("Date prévue de déblocage", tracking=True)
    echeance_date = fields.Date("Date d`échéance", tracking=True)
    echeance_fin_date = fields.Date("Date d`échéance", tracking=True)
    note = fields.Text(string='Description', tracking=True)
    partner_id = fields.Many2one('res.partner', string='Client/ Fournisseur')
    type_caution = fields.Selection([('cs', 'CS'),
                                     ('cbe', 'CBE'),
                                     ('cra', 'CRA')], string='Type')
    reference_credit = fields.Char(string='Référence du dossier banque', tracking=True)
    reference_interne = fields.Many2one('account.move', string='Référence interne (N. facture)', tracking=True)
    ref_interne = fields.Char(string='Référence interne (N. facture)')
    ref_cheque = fields.Char(string='N. Cheque')
    lib = fields.Char(string='Libellé')
    lib_pay = fields.Char(string='Termes de paiement')
    client = fields.Char(string='Client/ Fournisseur')
    user_id = fields.Many2one(
        'res.users', string='Order representative', index=True, tracking=True, readonly=True,
        default=lambda self: self.env.user, check_company=True)
    partner = fields.Many2one('res.partner', string='Client / Fournisseur', index=True, tracking=True)
    template_file = fields.Binary(string='Template', compute='_compute_template_file')
    template_name = fields.Char(string='template', default='Template Excel')
    import_file = fields.Binary(string='Importation')
    import_name = fields.Char(string='template', default='Fichier Excel')
    banque_id = fields.Many2one(
        'credit.banque', string='Banque', index=True, tracking=True, required=True,
                                         domain="[('has_autorisation', '=', True)]")
    type = fields.Many2one(
        'credit.type', string='Ligne de crédit', index=True, tracking=True, domain="[('id', 'in', type_ids)]")
    ligne_autorisation = fields.Many2one('credit.autorisation', string='Autorisation',
                                         domain="[('banque.id', '=', banque_id)]", required=True,
                                         ondelete='cascade')
    type_ids = fields.Many2many(
        'credit.type',
        string='Ligne de crédit',
        compute='_compute_type_ids',
        store=True
    )
    date_create = fields.Datetime(string='Date de création', required=True, index=True, copy=False,
                                  default=fields.Datetime.now,
                                  help="Indicates the date the operation deb was created.",
                                  readonly=True)
    disponible_id = fields.Many2one('credit.disponible', compute='compute_dispo', store=True)
    echeances = fields.One2many('credit.operation.deb.echeance', 'ref_opr_deb')
    taux = fields.Float(string='Taux', default=0)
    amount_invoice = fields.Float(string='Montant de la facture', default=0)
    file_ticket = fields.Binary(string='Document banque')
    file_name = fields.Char(string='File name')
    file_accord = fields.Binary(string='Accord de la banque')
    file_name1 = fields.Char(string='File name')
    remboursement = fields.Selection([('m', 'Mensuel'),
                                      ('t', 'Trimestriel'),
                                      ('s', 'Semestriel'),
                                      ])
    date_validite = fields.Date(string='Date de validité', related='ligne_autorisation.validite')
    taux_apply = fields.Float(string='Taux Appliqué')
    plm = fields.Float(string='PLM')
    type_id = fields.Integer(string='type', compute='compute_type', store=True)
    financement_hauteur = fields.Float(string='Financement à hauteur de %',
                                       related='ligne_autorisation.financement_hauteur')
    delai_mobilisation = fields.Integer(string='Délai de mobilisation',
                                        related='ligne_autorisation.delai_mobilisation')
    numero_traite = fields.Char(string='Numéro traite')
    devise = fields.Many2one('res.currency', string="Devise")
    montant_devise = fields.Float(string='Montant en devise')

    @api.depends('ligne_autorisation')
    def _compute_type_ids(self):
        for record in self:
            if record.ligne_autorisation:
                if record.ligne_autorisation.type_ids:
                    record.type_ids = record.ligne_autorisation.type_ids
                    record.type = record.ligne_autorisation.type_ids[0]
                else:
                    record.type_ids = [(5, 0, 0)]
            else:
                record.type_ids = [(5, 0, 0)]  # Clear the field if no ligne_autorisation

    @api.depends('type', 'amount_invoice',  'deblocage_date')
    def compute_type(self):
        for rec in self:
            if rec.type:
                rec.type_id = rec.type.id
            else:
                rec.type_id = 0
            if rec.deblocage_date and rec.delai_mobilisation:
                date_echeance = rec.deblocage_date + timedelta(days=rec.delai_mobilisation)
                rec.echeance_date = date_echeance
            if rec.financement_hauteur and rec.amount_invoice:
                rec.montant_debloque = rec.financement_hauteur * rec.amount_invoice

    @api.onchange('ligne_autorisation')
    def compute_dispo(self):
        for rec in self:
            dispo = self.env['credit.disponible'].search([('ligne_autorisation', '=', rec.ligne_autorisation.id)])
            rec.disponible_id = dispo

    def _compute_template_file(self):
        for rec in self:
            output = io.BytesIO()
            workbook = xlsxwriter.Workbook(output)
            worksheet = workbook.add_worksheet()
            # Ajouter les en-têtes de colonne
            headers = ['Montant à rembourser', 'Date d\'écheance']
            for col, header in enumerate(headers):
                worksheet.write(0, col, header)
            workbook.close()
            output.seek(0)

            # Convertir le fichier Excel en une chaîne d'octets et le sauvegarder dans le champ binaire
            rec.template_file = base64.b64encode(output.getvalue())

    def importExcel(self):
        for rec in self:
            if rec.import_file:
                wb = load_workbook(filename=io.BytesIO(base64.b64decode(rec.import_file)))
                ws = wb.active
                records_to_create = []
                # Iterate through rows and create records
                for row in ws.iter_rows(min_row=2, values_only=True):  # Assuming data starts from the second row
                    amount, due_date = row[:2]
                    record_vals = {
                        'montant_rembourser': amount,
                        'echeance_date': due_date,
                        'ref_opr_deb': rec.id
                    }
                    records_to_create.append(record_vals)
                # Create records using Odoo ORM
                for item in records_to_create:
                    self.env['credit.operation.deb.echeance'].create(item)

    @api.model
    def create(self, vals):
        if vals.get('name', _('New')) == _('New'):
            vals['name'] = self.env['ir.sequence'].next_by_code('credit.operation.deb') or _('New')
        if vals['deblocage_date'] and vals['echeance_date']:
            if self.echeance_date:
                if self.deblocage_date > self.echeance_date:
                    raise ValidationError(_("The echeance_date is less than the deblocage_date !!! "))
                else:
                    result = super(Operation_Deb, self).create(vals)
                    return result
            else:
                result = super(Operation_Deb, self).create(vals)
                return result
        else:
            result = super(Operation_Deb, self).create(vals)
            return result

    def write(self, vals):
        if vals.get('deblocage_date') == None:
            date_deb = self.deblocage_date
        else:
            date_deb = datetime.strptime(vals.get('deblocage_date'), '%Y-%m-%d').date()
        if vals.get('echeance_date') == None:
            date_ech = self.echeance_date
        elif type(vals.get('echeance_date'))==str:
            date_ech = datetime.strptime(vals.get('echeance_date'), '%Y-%m-%d').date()
        else:
            date_ech = vals['echeance_date']
        if date_deb and date_ech:
            if date_deb > date_ech:
                raise ValidationError(_("The echeance_date is less than the deblocage_date !!! "))
            else:
                for rec in self:
                    if vals.get('montant_debloque') == None:
                        mnt_debloc = rec.montant_debloque
                    else:
                        mnt_debloc = vals.get('montant_debloque')

                    if vals.get('montant_add') == None:
                        mnt_interet = rec.montant_add
                    else:
                        mnt_interet = vals.get('montant_add')
                    if mnt_interet != rec.montant_add or mnt_debloc != rec.montant_debloque:
                        disponible = rec.disponible_id
                        for debl in disponible.debloque_ids:
                            if debl.id == rec.id:
                                dispo = 0
                                total_autorisation = rec.disponible_id.montant_autorisation
                                total_deb = 0
                                total_ech = 0
                                if rec.disponible_id.has_deblocage:
                                    total_deb = sum(rec.disponible_id.debloque_ids.mapped('montant_rembourser'))
                                if rec.disponible_id.has_echeance:
                                    total_ech = sum(rec.disponible_id.echeance_ids.mapped('montant_a_rembourser'))
                                dispo = total_autorisation - total_deb + total_ech
                                rec.disponible_id.montant_disponible = dispo
                        for e in rec.echeances:
                            tmp = e.env['credit.echeance'].search([('echeance', '=', e.id)])
                            tmp.unlink()
                            p = e.env['credit.operation.p'].search([('echeance', '=', e.id)])
                            p.unlink()
                    if (len(rec.echeances) == 0) and (rec.type in [9,10,11,12]):
                        echeance = self.env['credit.echeance'].search([('ref_opr_deb', '=', rec.id)])
                        echeance.unlink()
                        pay = self.env['credit.operation.p'].search([('ref_opr_deb', '=', rec.id)])
                        pay.unlink()
        result = super(Operation_Deb, self).write(vals)
        return result

    def unlink(self):
        for rec in self:
            disponible = rec.env['credit.disponible'].search([('ligne_autorisation', '=', rec.ligne_autorisation.id)])
            recupere = disponible.montant_disponible + rec.montant_rembourser
            disponible.montant_disponible = recupere
        return super(Operation_Deb, self).unlink()


    def action_Confirme(self):
        for rec in self:
            disponible = rec.env['credit.disponible'].search([('ligne_autorisation', '=', rec.ligne_autorisation.id)])
            m_dispo = rec.ligne_autorisation.montant - rec.montant_rembourser
            if not rec.deblocage_date:
                raise UserError('Vous devriez saisir d\'abord la date de deblocage')
            if rec.montant_rembourser < rec.ligne_autorisation.montant and rec.type_id != 9:
                m_dispo = disponible.montant_disponible - rec.montant_rembourser
                if not disponible:
                    disponible.create({
                        'ligne_autorisation': rec.ligne_autorisation.id,
                        'banque': rec.banque_id.id,
                        'type': rec.type.id
                    })
                dispo = 0
                total_autorisation = rec.disponible_id.montant_autorisation
                total_deb = 0
                total_ech = 0
                if rec.disponible_id.has_deblocage:
                    total_deb = sum(rec.disponible_id.debloque_ids.mapped('montant_rembourser'))
                if rec.disponible_id.has_echeance:
                    total_ech = sum(rec.disponible_id.echeance_ids.mapped('montant_a_rembourser'))
                dispo = total_autorisation - total_deb + total_ech
                rec.disponible_id.montant_disponible = dispo

            if not rec.echeances:
                echeance = self.env['credit.echeance'].search([('ref_opr_deb', '=', rec.id)])
                if not echeance:
                    echeance.env['credit.echeance'].create({
                        'name': rec.name,
                        'ref_opr_deb': rec.id,
                        'banque': rec.banque_id.id,
                        'type': rec.type.id,
                        'echeance_date': rec.echeance_date,
                        'montant_a_rembourser': rec.montant_rembourser,
                        'fournisseur': rec.partner.id,
                        'facture': rec.reference_interne.id,
                    })
                p = self.env['credit.operation.p'].search([('ref_opr_deb', '=', rec.id)])
                if not p:
                    p.env['credit.operation.p'].create({
                        'ref_opr_deb': rec.id,
                        'banque': rec.banque_id.id,
                        'type': rec.type.id,
                        'montant_a_rembourser': rec.montant_rembourser,
                        'fournisseur': rec.partner.id,
                        'facture': rec.reference_interne.id,
                        'reference_dossier': rec.reference_credit,
                        'date_echeance': rec.echeance_date,
                        'date_deblocage': rec.deblocage_date,
                        'ligne_autorisation': rec.ligne_autorisation.id,
                    })
            else:
                for e in rec.echeances:
                    tmp = e.env['credit.echeance'].search([('echeance', '=', e.id)])
                    if not tmp:
                        e.env['credit.echeance'].create({
                            'ref_opr_deb': rec.id,
                            'banque': rec.banque_id.id,
                            'type': rec.type.id,
                            'echeance_date': e.echeance_date,
                            'montant_a_rembourser': e.montant_rembourser,
                            'fournisseur': rec.partner.id,
                            'facture': rec.reference_interne.id,
                            'echeance': e.id
                        })
                        e.env['credit.operation.p'].create({
                            'ref_opr_deb': rec.id,
                            'banque': rec.banque_id.id,
                            'type': rec.type.id,
                            'montant_a_rembourser': e.montant_rembourser,
                            'fournisseur': rec.partner.id,
                            'facture': rec.reference_interne.id,
                            'reference_dossier': rec.reference_credit,
                            'date_echeance': e.echeance_date,
                            'date_deblocage': rec.deblocage_date,
                            'ligne_autorisation': rec.ligne_autorisation.id,
                            'echeance': e.id
                        })
                        e.montant_total = e.montant_rembourser
                    else:
                        e.update()

            rec.state = 'confirmed'
            if rec.type_ligne == '1':
                view_id = self.env.ref('credit_bancaire.deblocage_wizard_form').id
                return {
                    'name': 'Information',
                    'type': 'ir.actions.act_window',
                    'view_mode': 'form',
                    'res_model': 'wizard.credit.deblocage',
                    'view_id': view_id,
                    'target': 'new',
                    'context': {'deblocage': rec.id},
                }
            else:
                return True

    def action_prolongation(self):
        for rec in self:
            if rec.echeance_date_new < rec.echeance_date_old:
                raise ValidationError(_("The new date is less than the old date !!! "))
            else:
                print(rec.echeance_date_new)
                rec.ref_opr_deb.echeance_date = rec.echeance_date_new
                ech = rec.env['credit.echeance'].search([('ref_opr_deb', '=', rec.ref_opr_deb.id),
                                                         ('echeance', '=', None)])
                ech.echeance_date = rec.echeance_date_new
                pay = rec.env['credit.operation.p'].search(
                    [('ref_opr_deb', '=', rec.ref_opr_deb.id), ('echeance', '=', None)])
                print(pay.montant_a_rembourser)
                print(pay.date_echeance)
                pay.date_echeance = rec.echeance_date_new
                '''rec.ligne_autorisation.validite_old = rec.ligne_autorisation.validite
                rec.ligne_autorisation.validite = rec.echeance_date_new
                rec.echeance_date_old = rec.ligne_autorisation.validite_old'''
                self.state = 'confirm'
                print('origin', rec.date_origin)



    '''@api.depends('banque', 'type')
    def _compute_autorisation(self):
        for rec in self:
            rec.ligne_autorisation = rec.env['credit.autorisation'].search(
                [('banque.name', '=', rec.banque_id.name), ('type.name', '=', rec.type.name)])
            print(rec.ligne_autorisation)'''

    @api.depends('montant_debloque', 'montant_add')
    def _compute_total(self):
        for rec in self:
            print('hi')
            if rec.type_ligne != '2':
                rec.montant_total = rec.montant_debloque + rec.montant_add
                rec.montant_total_comp = rec.montant_debloque + rec.montant_add
            else:
                rec.montant_total = rec.montant_debloque + rec.montant_add
                rec.montant_total_comp = rec.montant_debloque + rec.montant_add

    @api.depends('montant_total')
    def _compute_reste(self):
        for rec in self:
            qs = rec.env['credit.operation.p'].search([('ref_opr_deb.name', '=', rec.name)])
            rec.montant_rembourser = rec.montant_total
            print('qs', qs)
            for q in qs:
                rec.montant_rembourser = rec.montant_rembourser - q.montant_paye



class Echeance_detail(models.Model):
    _name = 'credit.operation.deb.echeance'
    _inherit = ["mail.thread", 'mail.activity.mixin']
    _description = "Ensemble des echeaances"

    name = fields.Char(string='Echeance partiel', required=True, copy=False, readonly=True,
                       default=lambda self: _('New'))

    montant_rembourser = fields.Float(string="Montant a rembourser", store=True, )
    montant_total = fields.Float(string="Montant total", store=True, readonly=True)
    echeance_date = fields.Date("Date d`échéance", tracking=True, required=True)
    ref_opr_deb = fields.Many2one('credit.operation.deb', string="reference debloque")
    @api.model
    def update(e):
            echeance = e.env['credit.echeance'].search([('echeance', '=', e.id)])
            if echeance:
                echeance.montant_a_rembourser =e.montant_rembourser
                echeance.echeance_date =e.echeance_date
            pay = e.env['credit.operation.p'].search([('echeance', '=', e.id)])
            if pay:
                pay.montant_a_rembourser =e.montant_rembourser
                pay.date_echeance =e.echeance_date

    def unlink(self):
        for rec in self:
            echeance = rec.env['credit.echeance'].search([('echeance', '=', rec.id)])
            echeance.unlink()
            pay = rec.env['credit.operation.p'].search([('echeance', '=', rec.id)])
            pay.unlink()
        return super(Echeance_detail, self).unlink()